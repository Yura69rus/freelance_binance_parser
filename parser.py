from selenium.webdriver import ActionChains
from selenium import webdriver
import openpyxl

LINK = 'https://www.binance.com/uk-UA/trade/EUR_USDT?layout=basic'
list_currents = []
list_prices = []

site = webdriver.Chrome()
site.get(LINK)

file = openpyxl.load_workbook("data.xlsx")
active_table = file.active

# ------- убрать всплывающее окно
actions = ActionChains(site)
actions.move_by_offset(10, 10).perform()
actions.click().perform()
# -------

for current in site.find_elements_by_xpath('//div[@class="item-symbol-text"]'):
    list_currents.append(current.text)

for price in site.find_elements_by_xpath('//div[@class="item item-price"]'):
    list_prices.append(price.text)

crnt_row = 1
for current in list_currents:
    sheet = active_table.cell(row=crnt_row, column=1)
    sheet.value = current
    crnt_row += 1

prc_row = 1
for price in list_prices:
    sheet = active_table.cell(row=prc_row, column=2)
    sheet.value = price
    prc_row += 1


file.save("data.xlsx")
